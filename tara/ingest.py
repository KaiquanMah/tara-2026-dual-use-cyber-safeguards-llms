import json
from pathlib import Path
from .models import Prompt

'''
Reads prompt datasets from file into list[Prompt].

Entry point: load_prompts(path) - auto-detects JSON vs Excel by file extension.

JSON handler (_load_json) supports 4 layouts:

  Keyed dict     {"ID-001": {prompt_id:..., prompt:...}, "ID-002": {...}}
  Array          [{prompt_id:..., prompt:...}, {...}]
  Legacy wrapper {"prompts": [{...}, {...}]}
  Single object  {prompt_id:..., prompt:...}

Field mapper (_dict_to_prompt) renames specific keys:
  prompt_id -> id,   prompt -> text
  All other fields pass through as-is. Entire original dict is preserved
  in Prompt.raw so no data is ever lost during ingestion.

Excel handler (_load_excel) reads headers from row 1, maps columns to
Prompt fields. Requires openpyxl.
'''

def load_prompts(path: str | Path) -> list[Prompt]:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Prompt file not found: {path}")
    if path.suffix == ".json":
        return _load_json(path)
    elif path.suffix in (".xlsx", ".xls"):
        return _load_excel(path)
    else:
        raise ValueError(f"Unsupported prompt format: {path.suffix}")


def _load_json(path: Path) -> list[Prompt]:
    with open(path) as f:
        data = json.load(f)

    if isinstance(data, list):
        return [_dict_to_prompt(p) for p in data]

    if isinstance(data, dict):
        has_prompts_key = "prompts" in data
        values_are_dicts = any(isinstance(v, dict) for v in data.values())
        if has_prompts_key:
            return [_dict_to_prompt(p) for p in data["prompts"]]
        elif values_are_dicts:
            return [_dict_to_prompt(v) for v in data.values()]
        else:
            return [_dict_to_prompt(data)]

    raise ValueError(f"Unexpected JSON structure in {path}")
    

def _dict_to_prompt(d: dict) -> Prompt:
    field_map = {
        "prompt_id": "id",
        "prompt": "text",
        "category": "category",
        "category_id": "category_id",
        "scenario_id": "scenario_id",
        "scenario_name": "scenario_name",
        "intent_class": "intent_class",
        "risk_level": "risk_level",
        "artifact_type": "artifact_type",
        "expected_behavior": "expected_behavior",
        "evaluation_focus": "evaluation_focus",
    }
    kwargs = {}
    for src_key, dst_key in field_map.items():
        if src_key in d:
            kwargs[dst_key] = d[src_key]
    kwargs.setdefault("id", d.get("prompt_id", ""))
    kwargs.setdefault("text", d.get("prompt", ""))
    kwargs["raw"] = d
    return Prompt(**{k: v for k, v in kwargs.items() if k in Prompt.__dataclass_fields__})


def _load_excel(path: Path) -> list[Prompt]:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel support: pip install openpyxl")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    prompts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        refs_raw = row_dict.get("references", "[]")
        if isinstance(refs_raw, str):
            refs = json.loads(refs_raw) if refs_raw.strip() else []
        else:
            refs = []
        prompts.append(Prompt(
            id=str(row_dict.get("id", "")),
            category=str(row_dict.get("category", "")),
            subdomain=str(row_dict.get("subdomain", "")),
            persona=str(row_dict.get("persona", "")),
            text=str(row_dict.get("text", "")),
            notes=str(row_dict.get("notes", "")),
            references=refs,
        ))
    return prompts

